import openpyxl as xl
import pandas as pd
import numpy as np

filename = '/Users/test/Desktop/Cwrtmalle Oct19.xlsx'
wb = xl.load_workbook(filename)
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet3']

feed = ['meal', 'blend', 'cake', 'nut', 'mag chloride', 'wynngold']
bedding = ['sawdust']
vet_and_med = ['vecoxan', 'bovikalc', 'bolus', 'scour', 'formalin']
repairs = ['gate', 'tine', 'drinker', 'trough']
personal = ['pounder', 'pedigree', 'bakers', 'joules', 'persil', 'butchers', 'fabric']
dairy_chemicals = ['ambic', 'peracetic', 'iodine', 'circulation', 'wynnsan', 'disinfectant']
catitem = 'Category'
outputcols = {
              'Line Total': 'sum',
              'Line Vat': 'sum',
              'Line Goods Value': 'sum'
              }
max_col = sheet1.max_column + 1

def myfunc(cell, category, itemvar):
    for i in category:
        if i.lower() in str(cell.value).lower():
            sheet1.cell(row, 11).value = itemvar

for row in range(2, sheet1.max_row + 1):
    cell = sheet1.cell(row, 4)
    myfunc(cell, feed, 'feed')
    myfunc(cell, bedding, 'bedding')
    myfunc(cell, vet_and_med, 'vet and med')
    myfunc(cell, repairs, 'repairs')
    myfunc(cell, personal, 'personal')
    myfunc(cell, dairy_chemicals, 'dairy chemicals')
    #myfunc(cell, sundries, 'sundries')
    if sheet1.cell(row, 11).value is None:
        sheet1.cell(row, max_col).value = 'sundries'
    sheet1.cell(1, max_col).value = catitem
wb.save('/Users/test/Desktop/invoicenew.xlsx')


wb = xl.load_workbook('/Users/test/Desktop/invoicenew.xlsx')


df = pd.read_excel('/Users/test/Desktop/invoicenew.xlsx')
data = df.groupby(catitem).agg(outputcols)
writer = pd.ExcelWriter('/Users/test/Desktop/invoicenew.xlsx', engine='openpyxl')
writer.book = wb
data.to_excel(writer, sheet_name='summary')
writer.save()
writer.close()















